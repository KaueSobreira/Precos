from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, View
from django.contrib import messages
from django.urls import reverse_lazy, reverse
from django.forms import modelformset_factory
from django.db import transaction
from django.http import HttpResponse
from decimal import Decimal, InvalidOperation
import openpyxl

from .models import TabelaFrete, RegraFreteMatriz, RegraFreteSimples, DescontoNotaVendedor, RegraFreteEspecial

# --- Tabela Frete ---

class RegrasMatrizTemplateView(View):
    """Gera e baixa o modelo XLSX para importação"""
    
    def get(self, request, tabela_pk):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Modelo Importação"

        # Cabeçalhos
        headers = ['peso_inicio', 'peso_fim', 'preco_inicio', 'preco_fim', 'score_inicio', 'score_fim', 'valor_frete', 'ordem', 'excedente (0 ou 1)']
        ws.append(headers)

        # Exemplo de dados
        # Exemplo normal
        example1 = [0.000, 1.000, 0.00, 100.00, None, None, 15.90, 1, 0]
        # Exemplo score
        example2 = [0.000, 1.000, None, None, 0, 10, 20.00, 1, 0]
        # Exemplo excedente
        example3 = [0.000, 1.000, 0.00, 100.00, None, None, 50.00, 1, 1]
        
        ws.append(example1)
        ws.append(example2)
        ws.append(example3)

        # Configurar resposta HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=modelo_importacao_regras.xlsx'
        
        wb.save(response)
        return response

class RegrasSimplesTemplateView(View):
    """Gera e baixa o modelo XLSX para importação de Regras Simples"""
    
    def get(self, request, tabela_pk):
        tabela = get_object_or_404(TabelaFrete, pk=tabela_pk)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Modelo Importação"

        # Cabeçalhos
        # A semântica muda (kg ou R$) mas os campos são os mesmos
        label_unidade = "kg" if tabela.tipo == 'peso' else "R$"
        headers = [f'inicio ({label_unidade})', f'fim ({label_unidade})', 'valor_frete (R$)', 'excedente (0 ou 1)']
        ws.append(headers)

        # Exemplo de dados
        example = [0.000, 1.000, 15.90, 0]
        ws.append(example)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=modelo_importacao_regras_simples.xlsx'
        
        wb.save(response)
        return response

class RegrasMatrizImportView(View):
    template_name = 'tabela_frete/import_form.html'

    def get(self, request, tabela_pk):
        tabela = get_object_or_404(TabelaFrete, pk=tabela_pk)
        return render(request, self.template_name, {'tabela': tabela})

    def post(self, request, tabela_pk):
        tabela = get_object_or_404(TabelaFrete, pk=tabela_pk)
        arquivo = request.FILES.get('arquivo')
        substituir = request.POST.get('substituir') == 'on'

        if not arquivo:
            messages.error(request, 'Nenhum arquivo enviado.')
            return redirect('regras_matriz_import', tabela_pk=tabela.pk)

        if not arquivo.name.endswith('.xlsx'):
            messages.error(request, 'O arquivo deve ser um Excel (.xlsx).')
            return redirect('regras_matriz_import', tabela_pk=tabela.pk)

        try:
            wb = openpyxl.load_workbook(arquivo, data_only=True)
            ws = wb.active

            regras_criadas = []
            
            with transaction.atomic():
                if substituir:
                    tabela.regras_matriz.all().delete()

                rows = list(ws.rows)
                if len(rows) > 0:
                    data_rows = rows[1:] # Ignorar cabeçalho
                else:
                    data_rows = []

                if not data_rows:
                     messages.warning(request, 'Arquivo vazio ou sem dados.')
                     return redirect('regras_matriz_import', tabela_pk=tabela.pk)

                for i, row in enumerate(data_rows):
                    # Formato esperado: peso_ini, peso_fim, preco_ini, preco_fim, score_ini, score_fim, valor, ordem, excedente
                    values = [cell.value for cell in row]

                    if not any(values): # Linha vazia
                        continue

                    def clean_decimal(val):
                        if val is None: return None
                        if isinstance(val, str):
                            val = val.strip().replace(',', '.')
                            if val == '': return None
                        return Decimal(str(val))
                    
                    def clean_int(val):
                        if val is None: return None
                        try:
                            return int(val)
                        except:
                            return None

                    try:
                        # Colunas: A=0, B=1, ...
                        peso_ini = clean_decimal(values[0])
                        peso_fim = clean_decimal(values[1])
                        preco_ini = clean_decimal(values[2])
                        preco_fim = clean_decimal(values[3])
                        score_ini = clean_int(values[4])
                        score_fim = clean_int(values[5])
                        valor_frete = clean_decimal(values[6])
                        ordem = clean_int(values[7]) or 0
                        
                        excedente = False
                        if len(values) > 8:
                            exc_val = values[8]
                            if exc_val in [1, '1', True, 'True', 'sim', 'Sim']:
                                excedente = True

                        if valor_frete is None:
                            continue 

                        regras_criadas.append(RegraFreteMatriz(
                            tabela=tabela,
                            peso_inicio=peso_ini,
                            peso_fim=peso_fim,
                            preco_inicio=preco_ini,
                            preco_fim=preco_fim,
                            score_inicio=score_ini,
                            score_fim=score_fim,
                            valor_frete=valor_frete,
                            ordem=ordem,
                            excedente=excedente
                        ))
                    except (ValueError, InvalidOperation, IndexError) as e:
                        print(f"Erro linha {i+2}: {e}")
                        continue

                if regras_criadas:
                    RegraFreteMatriz.objects.bulk_create(regras_criadas)
                    messages.success(request, f'{len(regras_criadas)} regras importadas com sucesso!')
                else:
                    messages.warning(request, 'Nenhuma regra válida encontrada no arquivo.')

        except Exception as e:
            messages.error(request, f'Erro ao processar arquivo: {str(e)}')
            return redirect('regras_matriz_import', tabela_pk=tabela.pk)

        return redirect('tabela_frete_detail', pk=tabela.pk)

class RegrasSimplesImportView(View):
    template_name = 'tabela_frete/import_form.html'

    def get(self, request, tabela_pk):
        tabela = get_object_or_404(TabelaFrete, pk=tabela_pk)
        return render(request, self.template_name, {'tabela': tabela})

    def post(self, request, tabela_pk):
        tabela = get_object_or_404(TabelaFrete, pk=tabela_pk)
        arquivo = request.FILES.get('arquivo')
        substituir = request.POST.get('substituir') == 'on'

        if not arquivo:
            messages.error(request, 'Nenhum arquivo enviado.')
            return redirect('regras_simples_import', tabela_pk=tabela.pk)

        if not arquivo.name.endswith('.xlsx'):
            messages.error(request, 'O arquivo deve ser um Excel (.xlsx).')
            return redirect('regras_simples_import', tabela_pk=tabela.pk)

        try:
            wb = openpyxl.load_workbook(arquivo, data_only=True)
            ws = wb.active

            regras_criadas = []
            
            with transaction.atomic():
                if substituir:
                    tabela.regras_simples.all().delete()

                rows = list(ws.rows)
                if len(rows) > 0:
                    data_rows = rows[1:] # Ignorar cabeçalho
                else:
                    data_rows = []

                if not data_rows:
                     messages.warning(request, 'Arquivo vazio ou sem dados.')
                     return redirect('regras_simples_import', tabela_pk=tabela.pk)

                for i, row in enumerate(data_rows):
                    # Formato esperado: inicio, fim, valor_frete, excedente
                    values = [cell.value for cell in row]
                    if not any(values): continue

                    def clean_decimal(val):
                        if val is None: return None
                        if isinstance(val, str):
                            val = val.strip().replace(',', '.')
                            if val == '': return None
                        return Decimal(str(val))

                    try:
                        inicio = clean_decimal(values[0])
                        fim = clean_decimal(values[1])
                        valor_frete = clean_decimal(values[2])
                        
                        excedente = False
                        if len(values) > 3:
                            exc_val = values[3]
                            if exc_val in [1, '1', True, 'True', 'sim', 'Sim']:
                                excedente = True
                        
                        if valor_frete is None: continue

                        regras_criadas.append(RegraFreteSimples(
                            tabela=tabela,
                            inicio=inicio,
                            fim=fim,
                            valor_frete=valor_frete,
                            excedente=excedente
                        ))
                    except (ValueError, InvalidOperation, IndexError) as e:
                        print(f"Erro linha {i+2}: {e}")
                        continue

                if regras_criadas:
                    RegraFreteSimples.objects.bulk_create(regras_criadas)
                    messages.success(request, f'{len(regras_criadas)} regras importadas com sucesso!')
                else:
                    messages.warning(request, 'Nenhuma regra válida encontrada no arquivo.')

        except Exception as e:
            messages.error(request, f'Erro ao processar arquivo: {str(e)}')
            return redirect('regras_simples_import', tabela_pk=tabela.pk)

        return redirect('tabela_frete_detail', pk=tabela.pk)

class RegrasMatrizBulkEditView(View):
    template_name = 'tabela_frete/regras_bulk_edit.html'

    def get(self, request, tabela_pk):
        tabela = get_object_or_404(TabelaFrete, pk=tabela_pk)
        
        # Debug
        regras_count = RegraFreteMatriz.objects.filter(tabela=tabela).count()
        print(f"--- DEBUG: Tabela {tabela.nome} (ID {tabela.pk}) has {regras_count} matrix rules. ---")

        RegraFormSet = modelformset_factory(
            RegraFreteMatriz,
            fields=['ordem', 'peso_inicio', 'peso_fim', 'preco_inicio', 'preco_fim', 'score_inicio', 'score_fim', 'valor_frete', 'ativo', 'excedente'],
            extra=0,
            can_delete=True
        )
        formset = RegraFormSet(queryset=RegraFreteMatriz.objects.filter(tabela=tabela).order_by('ordem', 'peso_inicio'))
        print(f"--- DEBUG: Formset instantiated with {len(formset)} forms. ---")
        return render(request, self.template_name, {'tabela': tabela, 'formset': formset})

    def post(self, request, tabela_pk):
        tabela = get_object_or_404(TabelaFrete, pk=tabela_pk)
        RegraFormSet = modelformset_factory(
            RegraFreteMatriz,
            fields=['ordem', 'peso_inicio', 'peso_fim', 'preco_inicio', 'preco_fim', 'score_inicio', 'score_fim', 'valor_frete', 'ativo', 'excedente'],
            extra=0,
            can_delete=True
        )
        # É importante passar o queryset no POST também para garantir a integridade dos IDs
        formset = RegraFormSet(request.POST, queryset=RegraFreteMatriz.objects.filter(tabela=tabela).order_by('ordem', 'peso_inicio'))
        
        if formset.is_valid():
            instances = formset.save(commit=False)
            for instance in instances:
                instance.tabela = tabela
                instance.save()
            for obj in formset.deleted_objects:
                obj.delete()
            messages.success(request, 'Regras atualizadas com sucesso!')
            return redirect('tabela_frete_detail', pk=tabela.pk)
        else:
            print("--- Formset Errors ---")
            print(formset.errors)
            print(formset.non_form_errors())
            messages.error(request, 'Erro ao salvar. Verifique os campos.')
        
        return render(request, self.template_name, {'tabela': tabela, 'formset': formset})

class TabelaFreteListView(ListView):
    model = TabelaFrete
    template_name = 'canais_vendas/tabela_frete_list.html'
    context_object_name = 'tabelas'

class TabelaFreteDetailView(DetailView):
    model = TabelaFrete
    template_name = 'canais_vendas/tabela_frete_detail.html'
    context_object_name = 'tabela'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.object.tipo in ['matriz', 'matriz_score']:
            context['regras'] = self.object.regras_matriz.all().order_by('excedente', 'ordem', 'peso_inicio')
        else:
            context['regras'] = self.object.regras_simples.all().order_by('excedente', 'inicio')
        
        context['regras_especiais'] = self.object.regras_especiais.all().order_by('ordem')
        context['descontos'] = self.object.descontos_nota.all().order_by('nota')
        context['canais'] = self.object.canais.all()
        return context

class TabelaFreteCreateView(CreateView):
    model = TabelaFrete
    template_name = 'canais_vendas/tabela_frete_form.html'
    fields = ['nome', 'tipo', 'descricao', 'ativo', 'suporta_nota_vendedor', 'adicionar_taxa_fixa', 'valor_taxa_fixa', 'usa_tabela_excedente', 'usar_preco_promocao']
    success_url = reverse_lazy('tabela_frete_list')

class TabelaFreteUpdateView(UpdateView):
    model = TabelaFrete
    template_name = 'canais_vendas/tabela_frete_form.html'
    fields = ['nome', 'tipo', 'descricao', 'ativo', 'suporta_nota_vendedor', 'adicionar_taxa_fixa', 'valor_taxa_fixa', 'usa_tabela_excedente', 'usar_preco_promocao']
    def get_success_url(self):
        return reverse('tabela_frete_detail', kwargs={'pk': self.object.pk})

class TabelaFreteDeleteView(DeleteView):
    model = TabelaFrete
    template_name = 'canais_vendas/tabela_frete_confirm_delete.html'
    success_url = reverse_lazy('tabela_frete_list')


# --- Regra Matriz ---

class RegraMatrizCreateView(CreateView):
    model = RegraFreteMatriz
    template_name = 'tabela_frete/regra_form.html'
    fields = ['ordem', 'peso_inicio', 'peso_fim', 'preco_inicio', 'preco_fim', 'score_inicio', 'score_fim', 'valor_frete', 'ativo', 'excedente']

    def dispatch(self, request, *args, **kwargs):
        self.tabela = get_object_or_404(TabelaFrete, pk=kwargs['tabela_pk'])
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        form.instance.tabela = self.tabela
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tabela'] = self.tabela
        context['titulo'] = "Nova Regra Matriz"
        return context

    def get_success_url(self):
        return reverse('tabela_frete_detail', kwargs={'pk': self.tabela.pk})

class RegraMatrizUpdateView(UpdateView):
    model = RegraFreteMatriz
    template_name = 'tabela_frete/regra_form.html'
    fields = ['ordem', 'peso_inicio', 'peso_fim', 'preco_inicio', 'preco_fim', 'score_inicio', 'score_fim', 'valor_frete', 'ativo', 'excedente']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tabela'] = self.object.tabela
        context['titulo'] = "Editar Regra Matriz"
        return context

    def get_success_url(self):
        return reverse('tabela_frete_detail', kwargs={'pk': self.object.tabela.pk})

class RegraMatrizDeleteView(DeleteView):
    model = RegraFreteMatriz
    template_name = 'tabela_frete/regra_confirm_delete.html'
    
    def get_success_url(self):
        return reverse('tabela_frete_detail', kwargs={'pk': self.object.tabela.pk})


# --- Regra Simples ---

class RegraSimplesCreateView(CreateView):
    model = RegraFreteSimples
    template_name = 'tabela_frete/regra_simples_form.html'
    fields = ['inicio', 'fim', 'valor_frete', 'excedente', 'ativo']

    def dispatch(self, request, *args, **kwargs):
        self.tabela = get_object_or_404(TabelaFrete, pk=kwargs['tabela_pk'])
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        form.instance.tabela = self.tabela
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tabela'] = self.tabela
        context['titulo'] = f"Nova Regra ({self.tabela.get_tipo_display()})"
        return context

    def get_success_url(self):
        return reverse('tabela_frete_detail', kwargs={'pk': self.tabela.pk})

class RegraSimplesUpdateView(UpdateView):
    model = RegraFreteSimples
    template_name = 'tabela_frete/regra_simples_form.html'
    fields = ['inicio', 'fim', 'valor_frete', 'excedente', 'ativo']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tabela'] = self.object.tabela
        context['titulo'] = "Editar Regra"
        return context

    def get_success_url(self):
        return reverse('tabela_frete_detail', kwargs={'pk': self.object.tabela.pk})

class RegraSimplesDeleteView(DeleteView):
    model = RegraFreteSimples
    template_name = 'tabela_frete/regra_confirm_delete.html'
    
    def get_success_url(self):
        return reverse('tabela_frete_detail', kwargs={'pk': self.object.tabela.pk})


# --- Desconto Nota ---

class DescontoCreateView(CreateView):
    model = DescontoNotaVendedor
    template_name = 'tabela_frete/desconto_form.html'
    fields = ['nota', 'percentual_desconto']

    def dispatch(self, request, *args, **kwargs):
        self.tabela = get_object_or_404(TabelaFrete, pk=kwargs['tabela_pk'])
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        form.instance.tabela = self.tabela
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tabela'] = self.tabela
        return context

    def get_success_url(self):
        return reverse('tabela_frete_detail', kwargs={'pk': self.tabela.pk})

class DescontoUpdateView(UpdateView):
    model = DescontoNotaVendedor
    template_name = 'tabela_frete/desconto_form.html'
    fields = ['nota', 'percentual_desconto']

    def get_success_url(self):
        return reverse('tabela_frete_detail', kwargs={'pk': self.object.tabela.pk})

class DescontoDeleteView(DeleteView):
    model = DescontoNotaVendedor
    template_name = 'tabela_frete/regra_confirm_delete.html'
    
    def get_success_url(self):
        return reverse('tabela_frete_detail', kwargs={'pk': self.object.tabela.pk})

# --- Regra Especial ---

class RegraEspecialCreateView(CreateView):
    model = RegraFreteEspecial
    template_name = 'tabela_frete/regra_especial_form.html'
    fields = ['ordem', 'largura_min', 'altura_min', 'profundidade_min', 'peso_min', 'valor_frete', 'ativo']

    def dispatch(self, request, *args, **kwargs):
        self.tabela = get_object_or_404(TabelaFrete, pk=kwargs['tabela_pk'])
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        form.instance.tabela = self.tabela
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tabela'] = self.tabela
        context['titulo'] = "Nova Regra Especial"
        return context

    def get_success_url(self):
        return reverse('tabela_frete_detail', kwargs={'pk': self.tabela.pk})

class RegraEspecialUpdateView(UpdateView):
    model = RegraFreteEspecial
    template_name = 'tabela_frete/regra_especial_form.html'
    fields = ['ordem', 'largura_min', 'altura_min', 'profundidade_min', 'peso_min', 'valor_frete', 'ativo']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tabela'] = self.object.tabela
        context['titulo'] = "Editar Regra Especial"
        return context

    def get_success_url(self):
        return reverse('tabela_frete_detail', kwargs={'pk': self.object.tabela.pk})

class RegraEspecialDeleteView(DeleteView):
    model = RegraFreteEspecial
    template_name = 'tabela_frete/regra_confirm_delete.html'
    
    def get_success_url(self):
        return reverse('tabela_frete_detail', kwargs={'pk': self.object.tabela.pk})